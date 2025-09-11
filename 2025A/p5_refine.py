from matplotlib import pyplot as plt
from numba import jit
import numpy as np
import matplotlib as mpl
import pygad
import openpyxl as px
import os
import warnings

from fix_parameters import *

warnings.filterwarnings("ignore", category=UserWarning)
mpl.rcParams["font.sans-serif"] = ["SimHei"]
mpl.rcParams["axes.unicode_minus"] = False


N_FY, N_S, N_M = 5, 3, 3
ACC = 1e-2


@jit  # 使用jit可以达到与C++相当的速度
def init_single(P, i):
    # 初始化单个无人机的投弹和爆炸位置
    p_drop = np.zeros((N_S, 3))
    p_exp = np.zeros((N_S, 3))
    t_drop = np.zeros((N_S,))
    t_exp = np.zeros((N_S,))
    r_fy0 = LOC_FY[i]
    arg_fy = P[0]
    rate_fy = P[1]
    v_fy = np.array([np.cos(arg_fy), np.sin(arg_fy), 0]) * rate_fy
    for j in range(N_S):
        dt_drop = P[2 + j]
        t_drop[j] = dt_drop
        if j > 0:
            t_drop[j] += t_drop[j - 1]
        dt_exp = P[2 + N_S + j]
        t_exp[j] = t_drop[j] + dt_exp
        p_drop[j] = r_fy0 + v_fy * t_drop[j]
        p_exp[j] = (
            p_drop[j] + v_fy * dt_exp + 1 / 2 * G_GRAVITY_VECTOR * dt_exp * dt_exp
        )

    return p_drop, p_exp, t_exp


@jit
def init(P):
    # 初始化所有无人机的投弹和爆炸位置
    p_drop = np.zeros((N_FY, N_S, 3))
    p_exp = np.zeros((N_FY, N_S, 3))
    t_drop = np.zeros((N_FY, N_S))
    t_exp = np.zeros((N_FY, N_S))
    for i in range(N_FY):
        r_fy0 = LOC_FY[i]
        arg_fy = P[i, 0]
        rate_fy = P[i, 1]
        v_fy = np.array([np.cos(arg_fy), np.sin(arg_fy), 0]) * rate_fy
        for j in range(N_S):
            dt_drop = P[i, 2 + j]
            t_drop[i, j] = dt_drop
            if j > 0:
                t_drop[i, j] += t_drop[i, j - 1]
            dt_exp = P[i, 2 + N_S + j]
            t_exp[i, j] = t_drop[i, j] + dt_exp
            p_drop[i, j] = r_fy0 + v_fy * t_drop[i, j]
            p_exp[i, j] = (
                p_drop[i, j]
                + v_fy * dt_exp
                + 1 / 2 * G_GRAVITY_VECTOR * dt_exp * dt_exp
            )

    return p_drop, p_exp, t_exp


@jit
def get_d2(t, r_s0, t_exp, k):
    d2_max = -np.inf
    r_m, r_s = LOC_M[k] + V_M_VECTOR[k] * t, r_s0 + V_SMOKE_VECTOR * (t - t_exp)
    for r_t in POINTS_REAL_TARGET:
        r1, r2 = r_m - r_t, r_s - r_t
        rc = np.dot(r1, r2)
        # 实际上需要两个端点的判断，但问题情境中烟雾不会到达目标点位置，遂只使用了一个判断。
        if rc > np.dot(r1, r1):
            d2 = np.dot(r1 - r2, r1 - r2)
        else:
            rs = np.cross(r1, r2)
            d2 = np.dot(rs, rs) / np.dot(r1, r1)
        d2_max = max(d2_max, d2)
    return d2_max


def plot_d2t(r_s0, t_exp, k, l=0.0, r=20.0):
    t = np.linspace(l, r, 500)
    d = np.array([get_d2(ti, r_s0, t_exp, k) for ti in t])
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    ax.plot(t, d)
    ax.set_xlabel("时间t/s")
    ax.set_ylabel("距离平方d2/m2")
    ax.grid()
    plt.show()


@jit
def get_unconditional_shielding_interval(r_s0, t_exp, k):
    a, b = 0, 60
    while abs(b - a) > ACC:
        c = b - (b - a) * GOLDEN_RATIO
        d = a + (b - a) * GOLDEN_RATIO
        d2c, d2d = get_d2(c, r_s0, t_exp, k), get_d2(d, r_s0, t_exp, k)
        if d2c < R_SMOKE2:
            m = c
            break
        if d2d < R_SMOKE2:
            m = d
            break
        if d2c < d2d:
            b = d
        else:
            a = c
    else:
        return (np.inf, -np.inf)
    l, r = a, m
    while r - l > 1e-6:
        m = (l + r) / 2
        if get_d2(m, r_s0, t_exp, k) < R_SMOKE2:
            r = m
        else:
            l = m
    interval_l = (l + r) / 2
    l, r = m, b
    while r - l > ACC:
        m = (l + r) / 2
        if get_d2(m, r_s0, t_exp, k) < R_SMOKE2:
            l = m
        else:
            r = m
    interval_r = (l + r) / 2
    return (interval_l, interval_r)


@jit
def calculate_union_length(intervals):
    size = intervals.shape[0]
    total_length = 0.0
    if size == 0:
        return total_length

    flat_starts = np.empty(size, dtype=np.float64)
    flat_ends = np.empty(size, dtype=np.float64)

    for i in range(size):
        s, e = intervals[i]
        flat_starts[i] = s
        flat_ends[i] = e

    sorted_indices = np.argsort(flat_starts)

    current_start = flat_starts[sorted_indices[0]]
    current_end = flat_ends[sorted_indices[0]]
    for k in range(1, size):
        idx = sorted_indices[k]
        s = flat_starts[idx]
        e = flat_ends[idx]

        if s <= current_end:
            if e > current_end:
                current_end = e
        else:
            if current_end > current_start:
                total_length += current_end - current_start
            current_start = s
            current_end = e
    if current_end > current_start:
        total_length += current_end - current_start

    return total_length


#
@jit
def get_single_single_shielding_interval(r_s0, t_exp, k):
    # 计算某个无人机的某个烟雾弹对某个导弹的遮挡时间
    # solves/s    ACC    N_FY, N_S, N_M    targets
    # 6900        1e-6   1,    1,   1        12
    l0, r0 = get_unconditional_shielding_interval(r_s0, t_exp, k)
    l = max(l0, t_exp)
    r = min(r0, t_exp + T_SMOKE)
    return (l, r)


@jit
def get_single_fy_shielding_intervals(p_exps, t_exps):
    # 计算一个无人机的所有烟雾弹对所有导弹的遮挡时间区间
    intervals = np.empty((N_S * N_M, 2))
    for j in range(N_S):
        for k in range(N_M):
            r_s0 = p_exps[j]
            t_exp = t_exps[j]
            intervals[j * N_M + k] = get_single_single_shielding_interval(
                r_s0, t_exp, k
            )
    return intervals


@jit
def get_single_fy_shielding_time(p_exp, t_exp):
    # 计算一个无人机的所有烟雾弹对所有导弹的遮挡时间总和
    # solves/s    ACC    N_FY, N_S, N_M    targets
    # 5900        1e-6   1,    1,   1       12
    # 370         1e-12  5,    3,   3       12
    # 625         1e-6   5,    3,   3       12
    # 1300        1e-2   5,    3,   3       12
    intervals = get_single_fy_shielding_intervals(p_exp, t_exp)
    return calculate_union_length(intervals)


@jit
def get_single_missile_shielding_intervals(p_exps, t_exps, k):
    # 计算一个导弹被所有无人机的所有烟雾弹遮挡的时间总和
    intervals = np.empty((N_FY * N_S, 2))
    for i in range(N_FY):
        for j in range(N_S):
            r_s0 = p_exps[i][j]
            t_exp = t_exps[i][j]
            intervals[i * N_S + j] = get_single_single_shielding_interval(
                r_s0, t_exp, k
            )
    return intervals


@jit
def get_single_missile_shielding_time(p_exps, t_exps, k):
    # 计算一个导弹被所有无人机的所有烟雾弹遮挡的时间总和
    intervals = get_single_missile_shielding_intervals(p_exps, t_exps, k)
    return calculate_union_length(intervals)


@jit
def get_all_missiles_shielding_time(p_exps, t_exps):
    # 计算所有导弹被所有无人机的所有烟雾弹遮挡的时间总和
    # solves/s    ACC    N_FY, N_S, N_M    targets
    # 72          1e-12  5,    3,   3        12
    # 130         1e-6   5,    3,   3        12
    # 141         1e-5   5,    3,   3        12
    # 167         1e-4   5,    3,   3        12
    # 202         1e-3   5,    3,   3        12
    # 238         1e-2   5,    3,   3        12
    # 333         1e-1   5,    3,   3        12
    # 410         1e0    5,    3,   3        12
    total_time = 0.0
    for k in range(N_M):
        total_time += get_single_missile_shielding_time(p_exps, t_exps, k)
    return total_time


@jit
def get_finally_shielding_time(P):
    # 计算所有导弹被所有无人机的所有烟雾弹遮挡的时间总和
    _, p_exps, t_exps = init(P)
    return get_all_missiles_shielding_time(p_exps, t_exps)


def plot_solvespeed():
    # 画出精度与求解速度的关系
    plt.figure()
    neglogaccs = [0, 1, 2, 3, 4, 5, 6, 12]
    solvespers = [410, 333, 238, 202, 167, 141, 130, 72]
    plt.plot(neglogaccs, solvespers, marker="o")
    plt.xlabel("精度负对数")
    plt.ylabel("每秒求解次数")
    plt.title("精度与求解速度关系")
    plt.show()


def process_res(P, do_print=True, do_plot=True, do_save=True):
    # fmt:off
    t_valid = np.full((N_FY, N_S), -1.0)
    idx_eff = np.full((N_FY, N_S), -1)
    p_drops, p_exps, t_exps = init(P)
    t_fys = np.zeros((N_FY,))
    t_total = get_all_missiles_shielding_time(p_exps, t_exps)
    for i in range(N_FY):
        t_fys[i] = get_single_fy_shielding_time(p_exps[i], t_exps[i])
        for j in range(N_S):
            for k in range(N_M):
                l, r = get_single_single_shielding_interval(p_exps[i, j], t_exps[i, j], k)
                if l < r:
                    t_valid[i][j] = r - l
                    idx_eff[i][j] = k
                else:
                    continue

    if do_print:
        for i in range(N_FY):
            print(f"无人机{i+1}信息: ")
            print(f"方向角: {P[i,0]/np.pi*180:.2f}°")
            print(f"无人机速度: {P[i,1]:.2f} m/s")
            for j in range(N_S):
                if t_valid[i][j] < 0:
                        continue
                k = idx_eff[i][j]
                print(f"烟幕弹{j+1}投放点: ({p_drops[i,j,0]:.2f}, {p_drops[i,j,1]:.2f}, {p_drops[i,j,2]:.2f})")
                print(f"烟幕弹{j+1}起爆点: ({p_exps[i,j,0]:.2f}, {p_exps[i,j,1]:.2f}, {p_exps[i,j,2]:.2f})")
                print(f"烟幕弹{j+1}干扰的导弹为{k+1}号导弹, 有效干扰时间为: {t_valid[i][j]:.4f}s")
            print(f"对所有导弹的干扰时间总和为: {t_fys[i]:.4f}s")
            print()
        print(f"所有导弹被烟雾弹遮挡的时间总和为: {t_total:.4f}s")
    
    if do_plot:
        def plot_point(ax, point, color, marker, label):
            ax.scatter(point[0], point[1], point[2], c=color, marker=marker, label=label)

        def plot_line_segment(ax, start, end, color, linestyle):
            ax.plot(
                [start[0], end[0]],
                [start[1], end[1]],
                [start[2], end[2]],
                c=color,
                linestyle=linestyle,
            )

        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, projection="3d")
        ax.set_xlabel("X (m)")
        ax.set_ylabel("Y (m)")
        ax.set_zlabel("Z (m)")
        plot_point(ax, LOC_REAL_TARGET, "r", "*", "真实目标")
        plot_point(ax, LOC_FAKE_TARGET, "g", "x", "假目标")
        for k in range(N_M):
            plot_point(ax, LOC_M[k], "black", "D", "导弹" if k == 0 else None)
            plot_line_segment(ax, LOC_M[k], LOC_FAKE_TARGET, "black", "--")
        for i in range(N_FY):
            plot_point(ax, LOC_FY[i], "b", "s", "无人机" if i == 0 else None)
            length = 4000
            v_fy_vec = np.array([np.cos(P[i, 0]), np.sin(P[i, 0]), 0]) * P[i, 1]
            loc_fy_end = LOC_FY[i] + v_fy_vec / np.linalg.norm(v_fy_vec) * length
            plot_line_segment(ax, LOC_FY[i], loc_fy_end, "b", "--")
            for j in range(N_S):
                if not t_valid[i][j] >= 0:
                    continue
                plot_point(ax, p_drops[i][j], "orange", "^", "投放点" if i == 0 and j == 0 else None)
                plot_point(ax, p_exps[i][j], "purple", "v", "起爆点" if i == 0 and j == 0 else None)
        plt.legend()
        plt.savefig("无人机与烟幕弹位置示意图.png")
        plt.show()
    
    if do_save:
        filename = "result3.xlsx"
        if os.path.exists(filename):
            os.remove(filename)
        wb = px.Workbook()
        heads = [
            "无人机编号",
            "无人机运动方向",
            "无人机运动速度 (m/s)",
            "烟幕干扰弹编号",
            "烟幕干扰弹投放点的x坐标 (m)",
            "烟幕干扰弹投放点的y坐标 (m)",
            "烟幕干扰弹投放点的z坐标 (m)",
            "烟幕干扰弹起爆点的x坐标 (m)",
            "烟幕干扰弹起爆点的y坐标 (m)",
            "烟幕干扰弹起爆点的z坐标 (m)",
            "有效干扰时长 (s)",
            "干扰的导弹编号",
        ]
        sheet = wb.create_sheet(title="sheet1", index=0)
        for j in range(len(heads)):
            sheet.cell(1, j + 1, value=heads[j])
        for i in range(15):
            r = i // N_S
            c = i % N_S
            sheet.cell(i + 2, 1, value=f"FY{r + 1}")
            sheet.cell(i + 2, 2, value=P[r, 0] / np.pi * 180)
            sheet.cell(i + 2, 3, value=P[r, 1])
            sheet.cell(i + 2, 4, value=c + 1)
            for j in range(3):
                sheet.cell(i + 2, 5 + j, value=p_drops[r][c][j])
                sheet.cell(i + 2, 8 + j, value=p_exps[r][c][j])
            if t_valid[r][c] >= 0:
                sheet.cell(i + 2, 11, value=t_valid[r][c])
            if idx_eff[r][c] >= 0:
                sheet.cell(i + 2, 12, value=idx_eff[r][c] + 1)

        wb.save("result3.xlsx")
        print("结果已保存至 result3.xlsx")
    # fmt:on


def optimize_GA(boundss, i, n_iters=100, n_samples=40):
    def fitness_func(ga_instance, sol, sol_idx):
        _, p_exps, t_exps = init_single(sol, i)
        return get_single_fy_shielding_time(p_exps, t_exps)

    def on_generation(ga_instance):
        print(
            f"\r{ga_instance.generations_completed}/{n_iters}",
            f"当前无人机对所有导弹干扰时间之和: {ga_instance.best_solution()[1]:.6f}",
            end="",
        )

    bounds = boundss[i]
    low = [l for l, _ in bounds]
    high = [h for _, h in bounds]
    gene_space = [{"low": l, "high": h} for l, h in bounds]

    ga_instance = pygad.GA(
        num_genes=len(bounds),
        num_generations=n_iters,
        sol_per_pop=n_samples,
        num_parents_mating=5,
        fitness_func=fitness_func,
        init_range_low=low,
        init_range_high=high,
        gene_space=gene_space,
        gene_type=float,
        on_generation=on_generation,
    )

    print(f"优化无人机 {i+1}")
    ga_instance.run()
    print()
    sol, gval, _ = ga_instance.best_solution()
    return sol, gval


def cyc_optimize_GA(boundss, n_iters=100, n_samples=40):
    # 对"单个无人机产生的总遮挡时间"的求解速度的预估值
    solves_per_second = 417
    est_time = (n_iters * n_samples * N_FY) / solves_per_second
    print(f"预估耗时: {est_time:.2f}秒")

    N_SINGLE = 2 + 2 * N_S
    P = np.zeros((N_FY, N_SINGLE))
    for _i in range(N_FY):
        sol, gval = optimize_GA(boundss, i=_i, n_iters=n_iters, n_samples=n_samples)
        for _j in range(N_SINGLE):
            P[_i, _j] = sol[_j]
    return P

def get_history_res(k):
    # 一些计算结果
    # fmt:off
    ress = [np.array([[  3.13522332, 139.83433237,   0.50862218,   2.63989217,
          2.1747692 ,   4.06564971,   5.21308998,   5.95432045],
       [  4.31075776, 121.78413684,   4.63810973,   1.21597306,
          3.54181038,   4.30462634,   6.55983939,   6.40413618],
       [  1.94054771,  89.34503688,  25.58522984,   4.89093704,
          1.80654354,  12.79722798,   6.57147299,   5.55883556],
       [  4.06673886, 126.5208231 ,   1.96464477,   2.49204807,
          2.13904719,   9.74215483,  11.93634094,  12.90889622],
       [  1.83239879,  97.26648858,  18.15264052,   2.49558253,
          2.00113802,   3.21929218,  12.05554315,  18.18393026]]),]
    # fmt:on
    return ress[k]

if __name__ == "__main__":
    # fmt:off
    # # 最宽松边界条件
    # boundss = [
    #     [(0, 2 * np.pi)] + [(70, 140)] + [(0, 5)] + [(1, 2)] * 2 + [(0, 10)] * 3,
    #     [(0, 2 * np.pi)] + [(70, 140)] + [(0, 20)] + [(1, 4)] * 2 + [(0, 20)] * 3,
    #     [(0, 2 * np.pi)] + [(70, 140)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
    #     [(0, 2 * np.pi)] + [(70, 140)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
    #     [(0, 2 * np.pi)] + [(70, 140)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
    # ]

    boundss = [
        [(np.pi - np.pi / 360, np.pi)] + [(139, 140)] + [(0.5, 1)] + [(1, 5)] * 2 + [(0, 10)] * 3,
        [(np.pi, 3 / 2 * np.pi)] + [(70, 140)] + [(0, 20)] + [(1, 5)] * 2 + [(0, 20)] * 3,
        [(np.pi / 2, np.pi)] + [(70, 100)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
        [(np.pi, 3 / 2 * np.pi)] + [(70, 140)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
        [(np.pi / 2, np.pi)] + [(70, 140)] + [(0, 30)] + [(1, 5)] * 2 + [(0, 20)] * 3,
    ]
    # fmt:on

    # 三个参数分别为：边界条件，迭代次数，每代样本数
    P = cyc_optimize_GA(boundss, n_iters=200, n_samples=60)
    
    # P = get_history_res(0)  # 使用历史计算结果进行展示
    
    plt.close("all")

    process_res(P, do_print=True, do_plot=True, do_save=True)
