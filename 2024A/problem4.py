from numpy import sqrt, log, pi, arctan2, sin, cos, arctan
from numba import jit
from matplotlib import pyplot as plt
import matplotlib as mpl
import numpy as np
from openpyxl.styles import Alignment, Font
import openpyxl as px

mpl.rcParams["font.sans-serif"] = ["SimHei"]
mpl.rcParams["axes.unicode_minus"] = False
plt.style.use("ggplot")


def setup():
    global V_HEAD, N_QUEUE, L0, L1, R, D, B, TH_MIN, S_MIN, BD2
    V_HEAD = 1
    N_QUEUE = 223
    L0, L1 = 2.86, 1.65
    R = 4.5
    D = 1.7
    B = D / 2 / pi
    TH_MIN = R / B
    S_MIN = B / 2 * (TH_MIN * sqrt(1 + TH_MIN**2) + log(TH_MIN + sqrt(1 + TH_MIN**2)))
    BD2 = B / 2

    global O1_X, O1_Y, O2_X, O2_Y, R1, R2, C1, C12, ALPHA_11, ALPHA_21
    # 已知盘入螺线和盘出螺线与掉头区域交界处的极角 TH_MIN
    # px,py 为两切点坐标，nx,ny 为两切点切线方向
    # O1X,O1Y 为圆弧圆心
    px, py = B * TH_MIN * cos(TH_MIN), B * TH_MIN * sin(TH_MIN)
    nx, ny = -B * cos(TH_MIN) + TH_MIN * B * sin(TH_MIN), -B * sin(
        TH_MIN
    ) - TH_MIN * B * cos(TH_MIN)
    nq = sqrt(nx * nx + ny * ny)
    nx, ny = nx / nq, ny / nq
    k1, k2 = 2, 1
    r = (px * px + py * py) / (nx * py - ny * px)
    R1, R2 = k1 / (k1 + k2) * r, k2 / (k1 + k2) * r

    O1_X, O1_Y = px + ny * R1, py - nx * R1
    O2_X, O2_Y = -px - ny * R2, -py + nx * R2
    mx, my = (k2 * O1_X + k1 * O2_X) / (k1 + k2), (k2 * O1_Y + k1 * O2_Y) / (k1 + k2)

    ALPHA_11 = arctan2(py - O1_Y, px - O1_X)
    ALPHA_12 = arctan2(my - O1_Y, mx - O1_X)
    ALPHA_21 = arctan2(O1_Y - my, O1_X - mx)
    ALPHA_22 = arctan2(-py - O2_Y, -px - O2_X)

    _BETA_1 = (ALPHA_11 - ALPHA_12) % (2 * pi)
    _BETA_2 = (ALPHA_22 - ALPHA_21) % (2 * pi)

    ALPHA_12 = ALPHA_11 - _BETA_1
    ALPHA_22 = ALPHA_21 + _BETA_2

    C1, C2 = R1 * _BETA_1, R2 * _BETA_2
    C12 = C1 + C2


@jit
def next_point(s_prev, x_prev, y_prev, length):
    # 返回下一个点的自然坐标，x坐标，y坐标，切向x分量，切向y分量
    s = s_prev - length
    while True:
        x, y, tx, ty = s_to_xy(s)
        L = (x - x_prev) ** 2 + (y - y_prev) ** 2 - length**2
        Lp = 2 * ((x - x_prev) * tx + (y - y_prev) * ty)
        e = L / Lp
        s_new = s - e

        if abs(e) < 1e-12:
            break
        s = s_new
    return s_new, x, y, tx, ty


@jit
def s_to_xy(s):
    def inver_s(s):
        # 螺线长的反函数，由自然坐标求极角
        s /= BD2
        th = sqrt(s)
        while True:
            th_new = th / 2 + (s - log(th + sqrt(1 + th**2))) / (2 * sqrt(1 + th**2))
            if th - th_new < 1e-12:
                break
            th = th_new
        return th_new

    # 返回 x,y 坐标，切向x分量，切向y分量
    if s < 0:
        s = S_MIN - s
        th = inver_s(s)
        return (
            B * th * cos(th),
            B * th * sin(th),
            -(cos(th) - th * sin(th)) / sqrt(1 + th**2),
            -(sin(th) + th * cos(th)) / sqrt(1 + th**2),
        )
    elif s > C12:
        s = s - C12 + S_MIN
        th = inver_s(s)
        return (
            -B * th * cos(th),
            -B * th * sin(th),
            -(cos(th) - th * sin(th)) / sqrt(1 + th**2),
            -(sin(th) + th * cos(th)) / sqrt(1 + th**2),
        )
    else:
        if s < C1:
            return (
                O1_X + R1 * cos(ALPHA_11 - s / R1),
                O1_Y + R1 * sin(ALPHA_11 - s / R1),
                sin(ALPHA_11 - s / R1),
                -cos(ALPHA_11 - s / R1),
            )
        else:
            s -= C1
            return (
                O2_X + R2 * cos(ALPHA_21 + s / R2),
                O2_Y + R2 * sin(ALPHA_21 + s / R2),
                -sin(ALPHA_21 + s / R2),
                cos(ALPHA_21 + s / R2),
            )


@jit
def queue(s):
    # 返回龙头位于s时，龙队各点信息
    ss, xs, ys = np.zeros(N_QUEUE + 1), np.zeros(N_QUEUE + 1), np.zeros(N_QUEUE + 1)
    txs, tys = np.zeros(N_QUEUE + 1), np.zeros(N_QUEUE + 1)
    ss[0] = s
    xs[0], ys[0], txs[0], tys[0] = s_to_xy(ss[0])
    ss[1], xs[1], ys[1], txs[1], tys[1] = next_point(ss[0], xs[0], ys[0], L0)
    for i in range(1, N_QUEUE):
        ss[i + 1], xs[i + 1], ys[i + 1], txs[i + 1], tys[i + 1] = next_point(
            ss[i], xs[i], ys[i], L1
        )
    vs = np.zeros(N_QUEUE + 1)
    vs[0] = V_HEAD
    for i in range(N_QUEUE):
        vs[i + 1] = (
            vs[i]
            * (txs[i] * (xs[i + 1] - xs[i]) + tys[i] * (ys[i + 1] - ys[i]))
            / (txs[i + 1] * (xs[i + 1] - xs[i]) + tys[i + 1] * (ys[i + 1] - ys[i]))
        )
    # 自然坐标，x坐标，y坐标，切线x方向，切线y方向，速度
    return xs, ys, vs


@jit
def max_v_when_s0(s):
    # 返回龙头位于s时的龙队最大速度
    xs, ys, vs = queue(s)
    return np.max(vs)


def plot_spiral(s):
    # 画出龙头位于s时的龙队
    xs, ys, vs = queue(s)

    def plot_circle(x, y, r, s=0, e=2 * pi):
        theta = np.linspace(s, e, 100)
        x1 = r * cos(theta) + x
        x2 = r * sin(theta) + y
        plt.plot(x1, x2)

    this = np.linspace(TH_MIN, 4 * TH_MIN, 10000)
    thos = np.linspace(TH_MIN, 4 * TH_MIN, 10000)
    xis, yis = B * this * cos(this), B * this * sin(this)
    xos, yos = -B * thos * cos(thos), -B * thos * sin(thos)

    plot_circle(O1_X, O1_Y, R1, ALPHA_11, ALPHA_11 - C1 / R1)
    plot_circle(O2_X, O2_Y, R2, ALPHA_21, ALPHA_21 + (C12 - C1) / R2)
    plt.title("龙队位置图")
    plt.plot(xis, yis, "-.", c="purple", label="盘入螺线")
    plt.plot(xos, yos, "--", c="red", label="盘出螺线")
    plt.plot(xs, ys, c="brown", label="龙队")
    plt.scatter(xs, ys, s=6, c="black", label="把手", zorder=3)
    plt.legend()
    plt.axis("equal")
    plt.savefig(f"p4-第{s / V_HEAD:.2f}s龙队位置图.svg")
    plt.clf()


def write_file():
    wb = px.Workbook()
    sheet1 = wb.create_sheet(title="位置", index=0)
    sheet2 = wb.create_sheet(title="速度", index=1)

    rls1 = (
        ["龙头x (m)", "龙头y (m)"]
        + [
            f"第{i}节龙身{axis} (m)"
            for i in range(1, N_QUEUE - 1)
            for axis in ["x", "y"]
        ]
        + ["龙尾x (m)", "龙尾y (m)"]
        + ["龙尾（后）x (m)", "龙尾（后）y (m)"]
    )
    rls2 = (
        ["龙头速度 (m/s)"]
        + [f"第{i}节龙身速度 (m/s)" for i in range(1, N_QUEUE - 1)]
        + ["龙尾速度 (m/s)", "龙尾（后）速度 (m/s)"]
    )
    cls = [f"{t}s" for t in range(-100, 101)]

    for i in range(len(rls1)):
        sheet1.cell(i + 2, 1, value=rls1[i])
    for i in range(len(rls2)):
        sheet2.cell(i + 2, 1, value=rls2[i])
    for i in range(len(cls)):
        sheet1.cell(1, i + 2, value=cls[i])
        sheet2.cell(1, i + 2, value=cls[i])
    for t in range(0, 201):
        print(f"处理 {t+1} 组数据", end="\r")
        s = V_HEAD * t
        xs, ys, vs = queue(s - 100)
        for i in range(len(xs)):
            sheet1.cell(2 * i + 2, t + 2, value=xs[i])
            sheet1.cell(2 * i + 2, t + 2).number_format = "0.000000"
        for i in range(len(ys)):
            sheet1.cell(2 * i + 3, t + 2, value=ys[i])
            sheet1.cell(2 * i + 3, t + 2).number_format = "0.000000"
        for i in range(len(vs)):
            sheet2.cell(i + 2, t + 2, value=vs[i])
            sheet2.cell(i + 2, t + 2).number_format = "0.000000"
    print()

    column_width = (14, 12)
    center_align = Alignment(horizontal="center", vertical="center")
    standard_font = Font(name="Times New Roman", size=10)
    for t in range(202):
        width = column_width[0] if t == 0 else column_width[1]
        col_letter = sheet1.cell(1, t + 1).column_letter
        sheet1.column_dimensions[col_letter].width = width
        sheet2.column_dimensions[col_letter].width = width
        for i in range(len(rls1) + 1):
            sheet1.cell(i + 1, t + 1).alignment = center_align
            sheet1.cell(i + 1, t + 1).font = standard_font
        for i in range(len(rls2) + 1):
            sheet2.cell(i + 1, t + 1).alignment = center_align
            sheet2.cell(i + 1, t + 1).font = standard_font

    wb.save("result4.xlsx")


setup()
write_file()
for t in [-100, 0, 100]:
    plot_spiral(t * V_HEAD)
