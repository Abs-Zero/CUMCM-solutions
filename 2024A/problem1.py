from numpy import pi, sin, cos, log, sqrt
import numpy as np
from numba import jit
from matplotlib import pyplot as plt
import matplotlib as mpl
from openpyxl.styles import Alignment, Font
import openpyxl as px


mpl.rcParams["font.sans-serif"] = ["SimHei"]
mpl.rcParams["axes.unicode_minus"] = False
plt.style.use("ggplot")

N_QUEUE = 223  # 龙队节数
L0, L1 = 2.86, 1.65  # 龙头与龙身长度
D = 0.55  # 螺距
B = D / 2 / pi  # 螺线方程系数
TH_MAX = 16 * 2 * pi  # 龙头初始极角
V_HEAD = 1  # 龙头速度
# 初始时刻，螺线由原点至龙头的曲线长度
S_MAX = B / 2 * (TH_MAX * sqrt(1 + TH_MAX**2) + log(TH_MAX + sqrt(1 + TH_MAX**2)))

BD2 = B / 2  # 常用系数，用于加速计算


@jit
def s_to_xy(s):
    def inver_s(s):
        # 螺线长的反函数，由自然坐标求极角
        s /= BD2
        th = sqrt(s)
        while True:
            th_new = th / 2 + (s - log(th + sqrt(1 + th**2))) / (2 * sqrt(1 + th**2))
            if th - th_new < 1e-12:
                break
            th = th_new
        return th_new

    s = S_MAX - s
    th = inver_s(s)
    return (
        B * th * cos(th),
        B * th * sin(th),
        -(cos(th) - th * sin(th)) / sqrt(1 + th**2),
        -(sin(th) + th * cos(th)) / sqrt(1 + th**2),
    )


@jit
def next_point(s_prev, x_prev, y_prev, length):
    # 返回下一个点的自然坐标，x坐标，y坐标，切向x分量，切向y分量
    s = s_prev - length
    while True:
        x, y, tx, ty = s_to_xy(s)
        d2 = (x - x_prev) ** 2 + (y - y_prev) ** 2
        L = d2 - length * sqrt(d2)
        Lp = 2 * ((x - x_prev) * tx + (y - y_prev) * ty)
        s_new = s - L / Lp

        if s - s_new < 1e-12:
            break
        s = s_new
    return s_new, x, y, tx, ty

@jit
def queue(s):
    # 返回龙头位于s时，龙队各点信息
    ss, xs, ys = [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1)
    txs, tys = [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1)
    ss[0] = s
    xs[0], ys[0], txs[0], tys[0] = s_to_xy(ss[0])
    ss[1], xs[1], ys[1], txs[1], tys[1] = next_point(ss[0], xs[0], ys[0], L0)
    for i in range(1, N_QUEUE):
        ss[i + 1], xs[i + 1], ys[i + 1], txs[i + 1], tys[i + 1] = next_point(
            ss[i], xs[i], ys[i], L1
        )

    vs = [0.0] * (N_QUEUE + 1)
    vs[0] = V_HEAD
    for i in range(N_QUEUE):
        vs[i + 1] = (
            vs[i]
            * (txs[i] * (xs[i + 1] - xs[i]) + tys[i] * (ys[i + 1] - ys[i]))
            / (txs[i + 1] * (xs[i + 1] - xs[i]) + tys[i + 1] * (ys[i + 1] - ys[i]))
        )
    return xs, ys, vs


def plot_spiral(s):
    # 画出龙头位于自然坐标 s 处时的龙队
    xs, ys, vs = queue(s)

    this = np.linspace(0, TH_MAX, 1000)
    xis, yis = B * this * cos(this), B * this * sin(this)
    plt.title("龙队位置图")
    plt.plot(xis, yis, "-.", c="purple", label="盘入螺线")
    plt.plot(xs, ys, c="brown", label="龙队")
    plt.scatter(xs, ys, s=6, c="black", label="把手", zorder=3)
    plt.legend()
    plt.axis("equal")
    plt.grid(False)
    plt.savefig(f"p1-龙头自然坐标为{s:.2f}时龙队位置图.svg")


def write_file():
    wb = px.Workbook()
    sheet1 = wb.create_sheet(title="位置", index=0)
    sheet2 = wb.create_sheet(title="速度", index=1)

    rls1 = (
        ["龙头x (m)", "龙头y (m)"]
        + [
            f"第{i}节龙身{axis} (m)"
            for i in range(1, N_QUEUE - 1)
            for axis in ["x", "y"]
        ]
        + ["龙尾x (m)", "龙尾y (m)"]
        + ["龙尾（后）x (m)", "龙尾（后）y (m)"]
    )
    rls2 = (
        ["龙头速度 (m/s)"]
        + [f"第{i}节龙身速度 (m/s)" for i in range(1, N_QUEUE - 1)]
        + ["龙尾速度 (m/s)", "龙尾（后）速度 (m/s)"]
    )
    cls = [f"{t}s" for t in range(301)]

    for i in range(len(rls1)):
        sheet1.cell(i + 2, 1, value=rls1[i])
    for i in range(len(rls2)):
        sheet2.cell(i + 2, 1, value=rls2[i])
    for i in range(len(cls)):
        sheet1.cell(1, i + 2, value=cls[i])
        sheet2.cell(1, i + 2, value=cls[i])
    for t in range(301):
        print(f"处理 {t} 组数据", end="\r")
        s = V_HEAD * t
        xs, ys, vs = queue(s)
        for i in range(len(xs)):
            sheet1.cell(2 * i + 2, t + 2, value=xs[i])
            sheet1.cell(2 * i + 2, t + 2).number_format = "0.000000"
        for i in range(len(ys)):
            sheet1.cell(2 * i + 3, t + 2, value=ys[i])
            sheet1.cell(2 * i + 3, t + 2).number_format = "0.000000"
        for i in range(len(vs)):
            sheet2.cell(i + 2, t + 2, value=vs[i])
            sheet2.cell(i + 2, t + 2).number_format = "0.000000"
    print()

    column_width = (14, 12)
    center_align = Alignment(horizontal="center", vertical="center")
    standard_font = Font(name="Times New Roman", size=10)
    for t in range(302):
        width = column_width[0] if t == 0 else column_width[1]
        col_letter = sheet1.cell(1, t + 1).column_letter
        sheet1.column_dimensions[col_letter].width = width
        sheet2.column_dimensions[col_letter].width = width
        for i in range(len(rls1) + 1):
            sheet1.cell(i + 1, t + 1).alignment = center_align
            sheet1.cell(i + 1, t + 1).font = standard_font
        for i in range(len(rls2) + 1):
            sheet2.cell(i + 1, t + 1).alignment = center_align
            sheet2.cell(i + 1, t + 1).font = standard_font

    wb.save("result1.xlsx")


write_file()  # 生成 Excel 文件
plot_spiral(150)  # 画出龙头位于自然坐标 s = 150 处时的龙队
