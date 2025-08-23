from numpy import pi, sin, cos, log, sqrt
import numpy as np
from numba import jit
from matplotlib import pyplot as plt
from matplotlib.patches import Polygon
import matplotlib as mpl
from openpyxl.styles import Alignment, Font
import openpyxl as px


mpl.rcParams["font.sans-serif"] = ["SimHei"]
mpl.rcParams["axes.unicode_minus"] = False
plt.style.use("ggplot")

N_QUEUE = 223  # 龙队节数
L0, L1 = 2.86, 1.65  # 龙头与龙身长度
D = 0.55  # 螺距
B = D / 2 / pi  # 螺线方程系数
TH_MAX = 16 * 2 * pi  # 龙头初始极角
V_HEAD = 1  # 龙头速度
# 初始时刻，螺线由原点至龙头的曲线长度
S_MAX = B / 2 * (TH_MAX * sqrt(1 + TH_MAX**2) + log(TH_MAX + sqrt(1 + TH_MAX**2)))

BD2 = B / 2  # 常用系数，用于加速计算

H1, H2 = 0.275, 0.15  # 把手到边缘的距离


@jit
def s_to_xy(s):
    def inver_s(s):
        # 螺线长的反函数，由自然坐标求极角
        s /= BD2
        th = sqrt(s)
        while True:
            th_new = th / 2 + (s - log(th + sqrt(1 + th**2))) / (2 * sqrt(1 + th**2))
            if th - th_new < 1e-12:
                break
            th = th_new
        return th_new

    s = S_MAX - s
    th = inver_s(s)
    return (
        B * th * cos(th),
        B * th * sin(th),
        -(cos(th) - th * sin(th)) / sqrt(1 + th**2),
        -(sin(th) + th * cos(th)) / sqrt(1 + th**2),
    )


@jit
def next_point(s_prev, x_prev, y_prev, length):
    # 返回下一个点的自然坐标，x坐标，y坐标，切向x分量，切向y分量
    s = s_prev - length
    while True:
        x, y, tx, ty = s_to_xy(s)
        d2 = (x - x_prev) ** 2 + (y - y_prev) ** 2
        L = d2 - length * sqrt(d2)
        Lp = 2 * ((x - x_prev) * tx + (y - y_prev) * ty)
        s_new = s - L / Lp

        if s - s_new < 1e-12:
            break
        s = s_new
    return s_new, x, y, tx, ty


@jit
def get_verts(p1_x, p1_y, p2_x, p2_y):
    # 计算把手p1与p2的四个顶点坐标
    vec_x, vec_y = p1_x - p2_x, p1_y - p2_y
    vec_q = sqrt(vec_x**2 + vec_y**2)
    n_x, n_y = vec_x / vec_q, vec_y / vec_q

    return (
        (p1_x + n_x * H1 + n_y * H2, p1_y + n_y * H1 - n_x * H2),  # 右前
        (p1_x + n_x * H1 - n_y * H2, p1_y + n_y * H1 + n_x * H2),  # 左前
        (p2_x - n_x * H1 - n_y * H2, p2_y - n_y * H1 + n_x * H2),  # 左后
        (p2_x - n_x * H1 + n_y * H2, p2_y - n_y * H1 - n_x * H2),  # 右后
    )


@jit
def queue(s):
    # 返回龙头位于s时，龙队各点信息
    ss, xs, ys = [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1)
    txs, tys = [0.0] * (N_QUEUE + 1), [0.0] * (N_QUEUE + 1)
    ss[0] = s
    xs[0], ys[0], txs[0], tys[0] = s_to_xy(ss[0])
    ss[1], xs[1], ys[1], txs[1], tys[1] = next_point(ss[0], xs[0], ys[0], L0)
    for i in range(1, N_QUEUE):
        ss[i + 1], xs[i + 1], ys[i + 1], txs[i + 1], tys[i + 1] = next_point(
            ss[i], xs[i], ys[i], L1
        )

    vs = [0.0] * (N_QUEUE + 1)
    vs[0] = V_HEAD
    for i in range(N_QUEUE):
        vs[i + 1] = (
            vs[i]
            * (txs[i] * (xs[i + 1] - xs[i]) + tys[i] * (ys[i + 1] - ys[i]))
            / (txs[i + 1] * (xs[i + 1] - xs[i]) + tys[i + 1] * (ys[i + 1] - ys[i]))
        )
    return xs, ys, vs


@jit
def min_distance(s):
    def point_to_segment_distance(p, a, b):
        # 计算点p到线段ab的最短距离，p在ab外侧时标记为inf
        px, py = p
        ax, ay = a
        bx, by = b

        abx = bx - ax
        aby = by - ay
        apx = px - ax
        apy = py - ay

        dot_product = apx * abx + apy * aby
        if dot_product <= 0:
            return float("inf")

        ab_len_sq = abx * abx + aby * aby
        t = dot_product / ab_len_sq
        if t >= 1:
            return float("inf")

        proj_x = ax + t * abx
        proj_y = ay + t * aby

        return sqrt((px - proj_x) ** 2 + (py - proj_y) ** 2)

    def vertices_to_polyline_distances(verts, polyline):
        # 计算顶点数组中每个点到折线的最短距离
        distances = []
        for v in verts:
            min_dist = float("inf")
            for i in range(len(polyline) - 1):
                a = polyline[i]
                b = polyline[i + 1]
                dist = point_to_segment_distance(v, a, b)
                if dist < min_dist:
                    min_dist = dist
            distances.append(min_dist)
        return np.array(distances)

    xs, ys, _ = queue(s)
    p0_x, p0_y = xs[0], ys[0]  # 龙头坐标
    p1_x, p1_y = xs[1], ys[1]  # 龙头后第一个把手
    p2_x, p2_y = xs[2], ys[2]  # 龙头后第二个把手
    vec10_x, vec10_y = p0_x - p1_x, p0_y - p1_y  # 龙头与龙头后第一个把手的连线方向
    vec21_x, vec21_y = p2_x - p1_x, p2_y - p1_y  # 第一个把手与第二个把手的连线方向

    vec_q_1 = sqrt(vec10_x**2 + vec10_y**2)
    vec_q_2 = sqrt(vec21_x**2 + vec21_y**2)
    n1_x, n1_y = vec10_x / vec_q_1, vec10_y / vec_q_1  # 单位化
    n2_x, n2_y = vec21_x / vec_q_2, vec21_y / vec_q_2

    vl0_x = p0_x + n1_x * H1 - n1_y * H2
    vl0_y = p0_y + n1_y * H1 + n1_x * H2  # 龙头左前方顶点位置
    vr0_x = p0_x + n1_x * H1 + n1_y * H2
    vr0_y = p0_y + n1_y * H1 - n1_x * H2  # 龙头右前方顶点位置

    vlf1_x = p1_x + n2_x * H1 - n2_y * H2
    vlf1_y = p1_y + n2_y * H1 + n2_x * H2  # 龙头后第一个把手左前顶点位置
    vrf1_x = p1_x + n2_x * H1 + n2_y * H2
    vrf1_y = p1_y + n2_y * H1 - n2_x * H2  # 龙头后第一个把手右前顶点位置

    vlr1_x = p1_x - n1_x * H1 - n1_y * H2
    vlr1_y = p1_y - n1_y * H1 + n1_x * H2  # 龙头后第一个把手左后顶点位置
    vrr1_x = p1_x - n1_x * H1 + n1_y * H2
    vrr1_y = p1_y - n1_y * H1 - n1_x * H2  # 龙头后第一个把手右后顶点位置

    verts = np.array(
        [
            [vl0_x, vl0_y],
            [vr0_x, vr0_y],
            [vlf1_x, vlf1_y],
            [vrf1_x, vrf1_y],
            [vlr1_x, vlr1_y],
            [vrr1_x, vrr1_y],
        ]
    )
    return min(vertices_to_polyline_distances(verts, np.array([xs[3:], ys[3:]]).T)) - H2


def plot_min_distance(sl, sr, nums):
    ss = np.linspace(sl, sr, nums)
    min_diss = np.zeros_like(ss)
    for i in range(len(ss)):
        min_diss[i] = min_distance(ss[i])
        print(i, end="\r")
    plt.plot(ss, min_diss, label="最小距离")
    plt.xlabel("自然坐标 s (m)")
    plt.ylabel("最小距离 (m)")
    plt.legend()
    plt.savefig(f"p2-顶点与板凳身最小距离随s变化关系图.svg")
    plt.clf()


def plot_spiral(s):
    # 龙头位于自然坐标 s 处时的龙队位置图
    xs, ys, _ = queue(s)
    this = np.linspace(0, TH_MAX, 1000)
    xis, yis = B * this * cos(this), B * this * sin(this)
    # plt.plot(xs, ys, c="brown", label="龙队")
    plt.title("龙队位置图")
    plt.grid(False)
    plt.axis("equal")
    
    for i in range(len(xs) - 1):
        vertices = get_verts(xs[i], ys[i], xs[i + 1], ys[i + 1])
        rect = Polygon(
            vertices, closed=True, facecolor="none", edgecolor="black", linewidth=0.4
        )
        plt.gca().add_patch(rect)
    plt.scatter(xs, ys, s=0.6, c="black", label="把手", zorder=3)
    plt.plot(xis, yis, "-.", linewidth=0.5, c="purple", label="盘入螺线")
    legend = plt.legend(loc='upper right', facecolor='white', edgecolor='black')
    legend.get_frame().set_alpha(1)
    plt.xlim(-3, 3)
    plt.ylim(-3, 3)

    plt.savefig(f"p2-龙头自然坐标为{s:.2f}时龙队位置图.svg")
    plt.clf()


def write_file(collision_moment):
    wb = px.Workbook()
    sheet1 = wb.create_sheet(index=0)

    rls = (
        ["龙头"]
        + [f"第{i}节龙身" for i in range(1, N_QUEUE - 1)]
        + ["龙尾", "龙尾（后）"]
    )
    cls = ["横坐标x (m)", "纵坐标y (m)", "速度 (m/s)"]

    for i in range(len(rls)):
        sheet1.cell(i + 2, 1, value=rls[i])
    for i in range(len(cls)):
        sheet1.cell(1, i + 2, value=cls[i])

    s = V_HEAD * collision_moment
    xs, ys, vs = queue(s)
    for i in range(len(xs)):
        sheet1.cell(i + 2, 2, value=xs[i])
        sheet1.cell(i + 2, 2).number_format = "0.000000"
    for i in range(len(ys)):
        sheet1.cell(i + 2, 3, value=ys[i])
        sheet1.cell(i + 2, 3).number_format = "0.000000"
    for i in range(len(vs)):
        sheet1.cell(i + 2, 4, value=vs[i])
        sheet1.cell(i + 2, 4).number_format = "0.000000"

    column_width = (14, 12)
    center_align = Alignment(horizontal="center", vertical="center")
    standard_font = Font(name="Times New Roman", size=10)
    for t in range(4):
        width = column_width[0] if t == 0 else column_width[1]
        col_letter = sheet1.cell(1, t + 1).column_letter
        sheet1.column_dimensions[col_letter].width = width
        for i in range(len(rls) + 1):
            sheet1.cell(i + 1, t + 1).alignment = center_align
            sheet1.cell(i + 1, t + 1).font = standard_font

    wb.save("result2.xlsx")


def solve(f, l, r, acc=1e-12):
    while r - l > acc:
        m = (l + r) / 2
        if f(m) * f(l) < 0:
            r = m
        else:
            l = m
    return (l + r) / 2


plot_min_distance(S_MAX - 40, S_MAX - 20, 200)
collision_moment = solve(min_distance, 412.4, 412.5) / V_HEAD
print(f"碰撞时刻: {collision_moment:.6f}s")
plot_spiral(collision_moment * V_HEAD)  # 碰撞时刻，龙队位置图
write_file(collision_moment)
